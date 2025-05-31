import csv
import os
import platform
from io import BytesIO, StringIO
from datetime import datetime
import oracledb
from flask import Flask, abort, make_response, render_template, request
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import DEFAULT_FONT, Font
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

from extrator_dgt import settings

#
# Configurar Instant Client - Oracle
#
cwd = os.path.dirname(os.path.abspath(__file__))

if platform.system() == "Windows":
    vendor_oracle_path = os.path.join(cwd, "vendor", "Oracle", "Windows")
    oracledb.init_oracle_client(lib_dir=vendor_oracle_path)
else:
    oracledb.init_oracle_client()


#
# App
#
app = Flask(__name__)

@app.route("/seplan_or_download/", methods=["GET"])
@app.route("/seplan_or_download", methods=["GET"])
def seplan_or_download():
    # Obter parâmetros de filtro da URL
    ano = request.args.get("ano", default="%")
    mes = request.args.get("mes", default="%")
    descricaoacao = request.args.get("descricaoacao", default="%")
    descricaoorcamentaria = request.args.get("descricaoorcamentaria", default="%")
    gnd = request.args.get("gnd", default="%")
    programaticaorcamentaria = request.args.get("programaticaorcamentaria", default="%")
    descricaoprograma = request.args.get("descricaoprograma", default="%")
    formato_csv = request.args.get("csv", default=None)

    try:
        # Conectar ao banco de dados
        connection = oracledb.connect(
            user="TRANSPARENCIATJBA",
            password="TRANSPARENCIATJBA",
            host="exacc03-stg-scan.tjba.jus.br",
            port=1521,
            service_name="tjprdhml",
        )

        # Criar um cursor
        cursor = connection.cursor()

        # Executar a consulta SQL com parâmetros de filtro
        query = """
            SELECT 
                COD_UNI_ORCAMENTARIA AS "Código",
                DESCRICAO_ORCAMENTARIA AS "Descrição",
                FUNC_ORCAMENTARIA AS "Função e Subfunção",
                PROGRAMATICA_ORCAMENTARIA AS "Programática (Programa, Ação e Subtítulo)",
                DESCRICAO_PROGRAMA AS "Programa",
                DESCRICAO_ACAO AS "Ação e Subtítulo",
                ESFERA_ORCAMENTARIA AS "Esfera",
                COD_FONTE_ORCAMENTARIA AS "Código",
                DESCRICAO_FONTE_ORCAMENTARIA AS "Descrição",
                GND,
                DOTACAO_INICIAL AS "Dotação Inicial",
                ACRESCIMOS_CRED_ADICIONAL AS "Acréscimos",
                DECRESCIMOS_CRED_ADICIONAL AS "Decréscimos",
                DOTACAO_ATUALIZADA AS "Dotação Atualizada",
                CONTINGENCIADO AS "Contingenciado",
                PROVISAO_MOV_LIQ_CRED AS "Provisão",
                DESTAQUE_MOV_LIQ_CRED AS "Destaque",
                DOTACAO_LIQUIDA AS "Dotação Líquida",
                EMPENHADO_EXECUCAO AS "Empenhado",
                EMPENHADO_EXECUCAO_PORCENTO AS "% Empenhado",
                LIQUIDADO_EXECUCAO AS "Liquidado",
                LIQUIDADO_EXECUCAO_PORCENTO AS "% Liquidado",
                PAGO_EXECUCAO AS "Pago",
                PAGO_EXECUCAO_PORCENTO AS "% Pago",
                MES_REFERENCIA,
                ANO_REFERENCIA
            FROM TRANSPARENCIATJBA.VW_INFO_ORCAMENTARIA_CONSOLIDADO
            WHERE (:ano IS NULL OR ANO_REFERENCIA LIKE :ano)
            AND (:mes IS NULL OR MES_REFERENCIA LIKE :mes)
            AND (:descricaoacao IS NULL OR DESCRICAO_ACAO LIKE :descricaoacao)
            AND (:descricaoorcamentaria IS NULL OR DESCRICAO_ORCAMENTARIA LIKE :descricaoorcamentaria)
            AND (:programaticaorcamentaria IS NULL OR PROGRAMATICA_ORCAMENTARIA LIKE :programaticaorcamentaria)
            AND (:gnd IS NULL OR GND LIKE :gnd)
            AND (:descricaoprograma IS NULL OR DESCRICAO_PROGRAMA LIKE :descricaoprograma)
            OFFSET 0 ROWS FETCH NEXT 100000 ROWS ONLY
        """
        cursor.execute(query, ano=ano, mes=mes,
                       descricaoacao=descricaoacao,
                       descricaoorcamentaria=descricaoorcamentaria,
                       gnd=gnd,
                       programaticaorcamentaria=programaticaorcamentaria,
                       descricaoprograma=descricaoprograma)

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

        if not formato_csv:
            template_path = os.path.join(app.root_path, 'templates', 'Template.xlsx')
            wb = load_workbook(template_path)
            ws = wb.active

            # --- 1. Modificação para a Célula B4 (Data de Referência) ---
            today_date = datetime.now().strftime("%d/%m/%Y")
            ws['B4'] = f"Data de referência: {today_date}"

            # --- 2. Desmesclar APENAS a área de dados ---
            data_start_row = 10
            data_start_col = 2  # Coluna B (índice 2)
            num_data_columns = len(columns)
            data_end_col_index = data_start_col + num_data_columns - 1

            max_rows_to_unmerge = data_start_row + len(rows) + 50 

            merged_ranges_to_unmerge_in_data_area = []
            for merged_range in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = merged_range.bounds
                if (min_row >= data_start_row and min_col <= data_end_col_index and max_col >= data_start_col) or \
                   (min_row < data_start_row and max_row >= data_start_row and min_col <= data_end_col_index and max_col >= data_start_col):
                    merged_ranges_to_unmerge_in_data_area.append(merged_range)

            for merged_range in merged_ranges_to_unmerge_in_data_area:
                ws.unmerge_cells(str(merged_range))

            # --- 3. Adicionar cópia de formatação para as linhas de dados ---
            reference_row_index_for_style = 9 # Linha 9 (cabeçalhos) para copiar bordas
            reference_cells_for_style = [ws.cell(row=reference_row_index_for_style, column=col_idx) 
                                         for col_idx in range(data_start_col, data_end_col_index + 1)]

            # Preencher os dados e aplicar formatação
            for r_idx, row_data in enumerate(rows):
                current_data_row = data_start_row + r_idx
                for c_idx, cell_value in enumerate(row_data):
                    target_cell = ws.cell(row=current_data_row, column=data_start_col + c_idx, value=cell_value)

                    if c_idx < len(reference_cells_for_style):
                        ref_cell = reference_cells_for_style[c_idx]

                        if ref_cell.font:
                            target_cell.font = ref_cell.font.copy()
                        if ref_cell.fill:
                            target_cell.fill = ref_cell.fill.copy()
                        if ref_cell.border:
                            target_cell.border = ref_cell.border.copy()
                        if ref_cell.alignment:
                            target_cell.alignment = ref_cell.alignment.copy()

            # --- NOVO: Adicionar Filtros (Triângulos) na Tabela ---
            # Definir o início da área de filtro (linha do cabeçalho da tabela de dados)
            filter_start_cell_col_letter = get_column_letter(data_start_col) # Coluna B
            filter_start_cell_row = reference_row_index_for_style # Linha 9 (onde estão os cabeçalhos para o filtro)

            # Definir o fim da área de filtro (última coluna de dados e última linha de dados)
            filter_end_cell_col_letter = get_column_letter(data_end_col_index)
            
            # Se houver dados, o filtro deve ir até a última linha de dados.
            # Caso contrário, o filtro deve ser aplicado apenas na linha do cabeçalho.
            if len(rows) > 0:
                filter_end_cell_row = data_start_row + len(rows) - 1
            else:
                filter_end_cell_row = reference_row_index_for_style # Apenas a linha do cabeçalho

            # Constrói a string do range do filtro, por exemplo "B9:AA100"
            filter_range = f"{filter_start_cell_col_letter}{filter_start_cell_row}:{filter_end_cell_col_letter}{filter_end_cell_row}"
            
            # Aplica o filtro automático na planilha
            ws.auto_filter.ref = filter_range


            xlsx_data = BytesIO()
            wb.save(xlsx_data)
            xlsx_data.seek(0)

            response = make_response(xlsx_data.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=DadosOrcamentoConsolidao.xlsx"
            response.headers["Content-type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            writer.writerows(rows)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=DadosOrcamentoConsolidao.csv"
            response.headers["Content-type"] = "text/csv"

        return response

    except Exception as e:
        abort(500, description=str(e))

    finally:
        # Certifique-se de fechar a conexão
        if "connection" in locals() and connection:
            connection.close()



@app.route("/seplan_or/", methods=["GET"])
@app.route("/seplan_or", methods=["GET"])
def seplan_or():
    return render_template("seplan_or.html")


@app.route("/", methods=["GET"])
def index():
    return "Extrator DGT"
