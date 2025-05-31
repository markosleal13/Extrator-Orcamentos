import csv
import os
import platform
from io import BytesIO, StringIO

import oracledb
from flask import Flask, abort, make_response, render_template, request
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import DEFAULT_FONT, Font
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

from extrator_dgt import settings

#
# Configurar Instant Client - Oracle
#
cwd = os.path.dirname(os.path.abspath(__file__))

if platform.system() == "Windows":
    vendor_oracle_path = os.path.join(cwd, "vendor", "Oracle", "Windows")
    oracledb.init_oracle_client(lib_dir=vendor_oracle_path)
else:
    oracledb.init_oracle_client()


#
# App
#
app = Flask(__name__)


@app.route("/ppp_download/", methods=["GET"])
@app.route("/ppp_download", methods=["GET"])
def ppp_download():
    # Obter parâmetros de filtro da URL
    comarca = request.args.get("comarca", default="%")
    vara = request.args.get("vara", default="%")
    classe = request.args.get("classe", default="%")
    localizacao = request.args.get("localizacao", default="%")
    situacao = request.args.get("situacao", default="%")
    assunto = request.args.get("assunto", default="%")
    paridade = request.args.get("paridade", default="%")
    natureza = request.args.get("natureza", default="%")
    formato_csv = request.args.get("csv", default=None)

    try:
        # Conectar ao banco de dados
        connection = oracledb.connect(
            user=settings.ppp_user,
            password=settings.ppp_password,
            host=settings.ppp_host,
            port=settings.ppp_port,
            service_name=settings.ppp_service_name,
        )

        # Criar um cursor
        cursor = connection.cursor()

        # Executar a consulta SQL com parâmetros de filtro
        query = """
            SELECT
                COMARCA || ' - ' || VARA AS "COMARCA - VARA",
                PROCESSO,
                PONTUACAO AS "PONTUAÇÃO",
                100 DIAS,
                TEMPO_MEDIO AS "TEMPO MÉDIO",
                CLASSE,
                NATUREZA,
                LOCALIZACAO_VALORES AS "LOCALIZAÇÃO VALORES",
                SITUACAO AS "SITUAÇÃO",
                CASE WHEN PARIDADE = 'I' THEN 'Ímpar' WHEN PARIDADE = 'P' THEN 'Par' ELSE PARIDADE END AS PARIDADE,
                CASE WHEN META_2 = 'S' THEN 'X' ELSE META_2 END AS "META 2",
                CASE WHEN META_4 = 'S' THEN 'X' ELSE META_4 END AS "META 4",
                CASE WHEN META_8 = 'S' THEN 'X' ELSE META_8 END AS "META 8",
                CASE WHEN META_10 = 'S' THEN 'X' ELSE META_10 END AS "META 10",
                CASE WHEN META_11 = 'S' THEN 'X' ELSE META_11 END AS "META 11",
                CASE WHEN VDO = 'S' THEN 'X' ELSE VDO END AS VDO,
                CASE WHEN FEM = 'S' THEN 'X' ELSE FEM END AS FEM,
                CASE WHEN MPU = 'S' THEN 'X' ELSE MPU END AS MPU,
                CASE WHEN SAUDE = 'S' THEN 'X' ELSE SAUDE END AS "SAÚDE",
                CASE WHEN ACAO_PENAL = 'S' THEN 'X' ELSE ACAO_PENAL END AS "AÇÃO PENAL",
                CASE WHEN AMBIENTAL = 'S' THEN 'X' ELSE AMBIENTAL END AS AMBIENTAL,
                CASE WHEN SANEAMENTO_12 = 'S' THEN 'X' ELSE SANEAMENTO_12 END AS "SANEAMENTO 12",
                CASE WHEN ADOCAO = 'S' THEN 'X' ELSE ADOCAO END AS "ADOÇÃO",
                ASSUNTO_PRINCIPAL AS "ASSUNTO PRINCIPAL"
            FROM UNIJUD.VW_PAINEL_PROCESSOS_PENDENTES
            WHERE (:comarca IS NULL OR COMARCA LIKE :comarca)
            AND (:vara IS NULL OR VARA LIKE :vara)
            AND (:classe IS NULL OR CLASSE LIKE :classe)
            AND (:localizacao IS NULL OR LOCALIZACAO_VALORES LIKE :localizacao)
            AND (:situacao IS NULL OR SITUACAO LIKE :situacao)
            AND (:assunto IS NULL OR ASSUNTO_PRINCIPAL LIKE :assunto)
            AND (:paridade IS NULL OR PARIDADE LIKE :paridade)
            AND (:natureza IS NULL OR NATUREZA LIKE :natureza)
            OFFSET 0 ROWS FETCH NEXT 100000 ROWS ONLY
        """
        cursor.execute(
            query,
            comarca=comarca,
            vara=vara,
            classe=classe,
            localizacao=localizacao,
            situacao=situacao,
            assunto=assunto,
            paridade=paridade,
            natureza=natureza,
        )

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

        if not formato_csv:
            # Criar um arquivo Excel com openpyxl
            wb = Workbook()
            ws = wb.active
            # Cabeçalho formatado
            header_font = Font(name="Arial", size=9, bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.append(columns)
            for col_num, _ in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center

            # Dados
            data_font = Font(name="Arial", size=8)
            for row in rows:
                ws.append(row)

            # Ajuste da largura das colunas
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Congelar cabeçalho
            ws.freeze_panes = "A2"

            # Filtro automático
            ws.auto_filter.ref = ws.dimensions


            xlsx_data = BytesIO()
            wb.save(xlsx_data)
            xlsx_data.seek(0)

            response = make_response(xlsx_data.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=dados.xlsx"
            response.headers["Content-type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            writer.writerows(rows)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=dados.csv"
            response.headers["Content-type"] = "text/csv"

        return response

    except Exception as e:
        abort(500, description=str(e))

    finally:
        # Certifique-se de fechar a conexão
        if "connection" in locals() and connection:
            connection.close()


# @app.route("/ppp_je_download/", methods=["GET"])
# @app.route("/ppp_je_download", methods=["GET"])
# def ppp_je_download():
#     # Obter parâmetros de filtro da URL
#     comarca = request.args.get("comarca", default="%")
#     vara = request.args.get("vara", default="%")
#     classe = request.args.get("classe", default="%")
#     localizacao = request.args.get("localizacao", default="%")
#     situacao = request.args.get("situacao", default="%")
#     assunto = request.args.get("assunto", default="%")
#     paridade = request.args.get("paridade", default="%")
#     natureza = request.args.get("natureza", default="%")
#     formato_csv = request.args.get("csv", default=None)

#     try:
#         # Conectar ao banco de dados
#         connection = oracledb.connect(
#             user=settings.ppp_user,
#             password=settings.ppp_password,
#             host=settings.ppp_host,
#             port=settings.ppp_port,
#             service_name=settings.ppp_service_name,
#         )

#         # Criar um cursor
#         cursor = connection.cursor()

#         # Executar a consulta SQL com parâmetros de filtro
#         query = """
#             SELECT
#                 COMARCA || ' - ' || VARA AS "COMARCA - VARA",
#                 PROCESSO,
#                 TEMPO_MEDIO AS "TEMPO MÉDIO",
#                 TEMPO_TRAMITACAO,
#                 LOCALIZACAO,
#                 NATUREZA,
#                 SITUACAO AS "SITUAÇÃO",
#                 CASE WHEN PARIDADE = 'I' THEN 'Ímpar' WHEN PARIDADE = 'P' THEN 'Par' ELSE PARIDADE END AS PARIDADE,
#                 CASE WHEN META_2 = 'S' THEN 'X' ELSE META_2 END AS "META 2",
#                 CASE WHEN META_4 = 'S' THEN 'X' ELSE META_4 END AS "META 4",
#                 CASE WHEN META_10 = 'S' THEN 'X' ELSE META_10 END AS "META 10",
#                 CASE WHEN SAUDE = 'S' THEN 'X' ELSE SAUDE END AS "SAÚDE",
#                 CASE WHEN ACAO_PENAL = 'S' THEN 'X' ELSE ACAO_PENAL END AS "AÇÃO PENAL",
#                 CASE WHEN AMBIENTAL = 'S' THEN 'X' ELSE AMBIENTAL END AS AMBIENTAL,
#                 CASE WHEN SANEAMENTO_12 = 'S' THEN 'X' ELSE SANEAMENTO_12 END AS "SANEAMENTO",
#                 CASE WHEN ADOCAO = 'S' THEN 'X' ELSE ADOCAO END AS "ADOÇÃO",
#                 ASSUNTO_PRINCIPAL AS "ASSUNTO PRINCIPAL"
#             FROM UNIJUD.VW_PAINEL_PROCESSOS_PENDENTES_JE
#             WHERE (:comarca IS NULL OR COMARCA LIKE :comarca)
#             AND (:vara IS NULL OR VARA LIKE :vara)
#             AND (:classe IS NULL OR CLASSE LIKE :classe)
#             AND (:localizacao IS NULL OR LOCALIZACAO_VALORES LIKE :localizacao)
#             AND (:situacao IS NULL OR SITUACAO LIKE :situacao)
#             AND (:assunto IS NULL OR ASSUNTO_PRINCIPAL LIKE :assunto)
#             AND (:paridade IS NULL OR PARIDADE LIKE :paridade)
#             AND (:natureza IS NULL OR NATUREZA LIKE :natureza)
#             OFFSET 0 ROWS FETCH NEXT 100000 ROWS ONLY
#         """
#         cursor.execute(
#             query,
#             comarca=comarca,
#             vara=vara,
#             classe=classe,
#             localizacao=localizacao,
#             situacao=situacao,
#             assunto=assunto,
#             paridade=paridade,
#             natureza=natureza,
#         )

#         rows = cursor.fetchall()
#         columns = [col[0] for col in cursor.description]

#         if not formato_csv:
#             # Criar um arquivo Excel com openpyxl
#             wb = Workbook()
#             ws = wb.active
#             ws.append(columns)

#             for row in rows:
#                 ws.append(row)

#             ws.auto_filter.ref = ws.dimensions
#             font = Font(name="Arial", sz=8)
#             {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}

#             xlsx_data = BytesIO()
#             wb.save(xlsx_data)
#             xlsx_data.seek(0)

#             response = make_response(xlsx_data.getvalue())
#             response.headers["Content-Disposition"] = "attachment; filename=dados.xlsx"
#             response.headers["Content-type"] = (
#                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
#         else:
#             output = StringIO()
#             writer = csv.writer(output)
#             writer.writerow(columns)
#             writer.writerows(rows)

#             response = make_response(output.getvalue())
#             response.headers["Content-Disposition"] = "attachment; filename=dados.csv"
#             response.headers["Content-type"] = "text/csv"

#         return response

#     except Exception as e:
#         abort(500, description=str(e))

#     finally:
#         # Certifique-se de fechar a conexão
#         if "connection" in locals() and connection:
#             connection.close()


@app.route("/ppp_tr_download/", methods=["GET"])
@app.route("/ppp_tr_download", methods=["GET"])
def ppp_tr_download():
    # Obter parâmetros de filtro da URL
    comarca = request.args.get("comarca", default="%")
    vara = request.args.get("vara", default="%")
    classe = request.args.get("classe", default="%")
    localizacao = request.args.get("localizacao", default="%")
    situacao = request.args.get("situacao", default="%")
    assunto = request.args.get("assunto", default="%")
    paridade = request.args.get("paridade", default="%")
    natureza = request.args.get("natureza", default="%")
    formato_csv = request.args.get("csv", default=None)

    try:
        # Conectar ao banco de dados
        connection = oracledb.connect(
            user=settings.ppp_user,
            password=settings.ppp_password,
            host=settings.ppp_host,
            port=settings.ppp_port,
            service_name=settings.ppp_service_name,
        )

        # Criar um cursor
        cursor = connection.cursor()

        # Executar a consulta SQL com parâmetros de filtro
        query = """
            SELECT
                COMARCA || ' - ' || VARA AS "COMARCA - VARA",
                PROCESSO,
                COD_PROCESSO AS "CÓDIGO DO PROCESSO",
                TEMPO_MEDIO AS "TEMPO MÉDIO",
                TEMPO_TRAMITACAO,
                NATUREZA,
                -- LOCALIZACAO,
                SITUACAO AS "SITUAÇÃO",
                CASE WHEN PARIDADE = 'I' THEN 'Ímpar' WHEN PARIDADE = 'P' THEN 'Par' ELSE PARIDADE END AS PARIDADE,
                CASE WHEN META_2 = 'S' THEN 'X' ELSE META_2 END AS "META 2",
                CASE WHEN META_4 = 'S' THEN 'X' ELSE META_4 END AS "META 4",
                CASE WHEN META_10 = 'S' THEN 'X' ELSE META_10 END AS "META 10",
                CASE WHEN SAUDE = 'S' THEN 'X' ELSE SAUDE END AS "SAÚDE",
                CASE WHEN ACAO_PENAL = 'S' THEN 'X' ELSE ACAO_PENAL END AS "AÇÃO PENAL",
                CASE WHEN AMBIENTAL = 'S' THEN 'X' ELSE AMBIENTAL END AS AMBIENTAL,
                CASE WHEN SANEAMENTO_12 = 'S' THEN 'X' ELSE SANEAMENTO_12 END AS "SANEAMENTO",
                CASE WHEN ADOCAO = 'S' THEN 'X' ELSE ADOCAO END AS "ADOÇÃO",
                ASSUNTO_PRINCIPAL AS "ASSUNTO PRINCIPAL"
            FROM UNIJUD.VW_PAINEL_PROCESSOS_PENDENTES_TR
            WHERE (:comarca IS NULL OR COMARCA LIKE :comarca)
            AND (:vara IS NULL OR VARA LIKE :vara)
            AND (:classe IS NULL OR CLASSE LIKE :classe)
            AND (:localizacao IS NULL OR LOCALIZACAO_VALORES LIKE :localizacao)
            AND (:situacao IS NULL OR SITUACAO LIKE :situacao)
            AND (:assunto IS NULL OR ASSUNTO_PRINCIPAL LIKE :assunto)
            AND (:paridade IS NULL OR PARIDADE LIKE :paridade)
            AND (:natureza IS NULL OR NATUREZA LIKE :natureza)
            OFFSET 0 ROWS FETCH NEXT 100000 ROWS ONLY
        """
        cursor.execute(
            query,
            comarca=comarca,
            vara=vara,
            classe=classe,
            localizacao=localizacao,
            situacao=situacao,
            assunto=assunto,
            paridade=paridade,
            natureza=natureza,

        )

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

        if not formato_csv:
            # Criar um arquivo Excel com openpyxl
            wb = Workbook()
            ws = wb.active
            ws.append(columns)

            for row in rows:
                ws.append(row)

            ws.auto_filter.ref = ws.dimensions
            font = Font(name="Arial", sz=8)
            {k: setattr(DEFAULT_FONT, k, v) for k, v in font.__dict__.items()}

            xlsx_data = BytesIO()
            wb.save(xlsx_data)
            xlsx_data.seek(0)

            response = make_response(xlsx_data.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=dados.xlsx"
            response.headers["Content-type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            writer.writerows(rows)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=dados.csv"
            response.headers["Content-type"] = "text/csv"

        return response

    except Exception as e:
        abort(500, description=str(e))

    finally:
        # Certifique-se de fechar a conexão
        if "connection" in locals() and connection:
            connection.close()

@app.route("/seplan_or_download/", methods=["GET"])
@app.route("/seplan_or_download", methods=["GET"])
def seplan_or_download():
    # Obter parâmetros de filtro da URL
    ano = request.args.get("ano", default="%")
    mes = request.args.get("mes", default="%")
    descricaoacao = request.args.get("descricaoacao", default="%")
    descricaoorcamentaria = request.args.get("descricaoorcamentaria", default="%")
    gnd = request.args.get("gnd", default="%")
    programaticaorcamentaria = request.args.get("programaticaorcamentaria", default="%")
    descricaoprograma = request.args.get("descricaoprograma", default="%")
    formato_csv = request.args.get("csv", default=None)

    try:
        # Conectar ao banco de dados
        connection = oracledb.connect(
            user="TRANSPARENCIATJBA",
            password="TRANSPARENCIATJBA",
            host="exacc03-stg-scan.tjba.jus.br",
            port=1521,
            service_name="tjprdhml",
        )

        # Criar um cursor
        cursor = connection.cursor()

        # Executar a consulta SQL com parâmetros de filtro
        query = """
            SELECT 
                COD_UNI_ORCAMENTARIA AS "Código",
                DESCRICAO_ORCAMENTARIA AS "Descrição",
                FUNC_ORCAMENTARIA AS "Função e Subfunção",
                PROGRAMATICA_ORCAMENTARIA AS "Programática (Programa, Ação e Subtítulo)",
                DESCRICAO_PROGRAMA AS "Programa",
                DESCRICAO_ACAO AS "Ação e Subtítulo",
                ESFERA_ORCAMENTARIA AS "Esfera",
                COD_FONTE_ORCAMENTARIA AS "Código",
                DESCRICAO_FONTE_ORCAMENTARIA AS "Descrição",
                GND,
                DOTACAO_INICIAL AS "Dotação Inicial",
                ACRESCIMOS_CRED_ADICIONAL AS "Acréscimos",
                DECRESCIMOS_CRED_ADICIONAL AS "Decréscimos",
                DOTACAO_ATUALIZADA AS "Dotação Atualizada",
                CONTINGENCIADO AS "Contingenciado",
                PROVISAO_MOV_LIQ_CRED AS "Provisão",
                DESTAQUE_MOV_LIQ_CRED AS "Destaque",
                DOTACAO_LIQUIDA AS "Dotação Líquida",
                EMPENHADO_EXECUCAO AS "Empenhado",
                EMPENHADO_EXECUCAO_PORCENTO AS "% Empenhado",
                LIQUIDADO_EXECUCAO AS "Liquidado",
                LIQUIDADO_EXECUCAO_PORCENTO AS "% Liquidado",
                PAGO_EXECUCAO AS "Pago",
                PAGO_EXECUCAO_PORCENTO AS "% Pago",
                MES_REFERENCIA,
                ANO_REFERENCIA
            FROM TRANSPARENCIATJBA.VW_INFO_ORCAMENTARIA_CONSOLIDADO
            WHERE (:ano IS NULL OR ANO_REFERENCIA LIKE :ano)
            AND (:mes IS NULL OR MES_REFERENCIA LIKE :mes)
            AND (:descricaoacao IS NULL OR DESCRICAO_ACAO LIKE :descricaoacao)
            AND (:descricaoorcamentaria IS NULL OR DESCRICAO_ORCAMENTARIA LIKE :descricaoorcamentaria)
            AND (:programaticaorcamentaria IS NULL OR PROGRAMATICA_ORCAMENTARIA LIKE :programaticaorcamentaria)
            AND (:gnd IS NULL OR GND LIKE :gnd)
            AND (:descricaoprograma IS NULL OR DESCRICAO_PROGRAMA LIKE :descricaoprograma)
            OFFSET 0 ROWS FETCH NEXT 100000 ROWS ONLY
        """
        cursor.execute(query, ano=ano, mes=mes,
                       descricaoacao=descricaoacao,
                       descricaoorcamentaria=descricaoorcamentaria,
                       gnd=gnd,
                       programaticaorcamentaria=programaticaorcamentaria,
                       descricaoprograma=descricaoprograma)

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

        if not formato_csv:
            template_path = os.path.join(app.root_path, 'templates', 'Template.xlsx')
            wb = load_workbook(template_path)
            ws = wb.active

            # --- INÍCIO DA MODIFICAÇÃO ---
            # Desmesclar todas as células da planilha antes de gravar os dados
            # Isso é crucial para evitar erros 'MergedCell' object attribute 'value' is read-only
            # ao tentar gravar em células que fazem parte de uma região mesclada.
            merged_ranges_to_remove = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges_to_remove:
                ws.unmerge_cells(str(merged_range))
            # --- FIM DA MODIFICAÇÃO ---


            start_row = 10
            start_col = 2 # Coluna B

            # Preencher os dados
            for r_idx, row_data in enumerate(rows):
                for c_idx, cell_value in enumerate(row_data):
                    ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=cell_value)

            xlsx_data = BytesIO()
            wb.save(xlsx_data)
            xlsx_data.seek(0)

            response = make_response(xlsx_data.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=DadosOrcamentoConsolidao.xlsx"
            response.headers["Content-type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            writer.writerows(rows)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=DadosOrcamentoConsolidao.csv"
            response.headers["Content-type"] = "text/csv"

        return response

    except Exception as e:
        abort(500, description=str(e))

    finally:
        # Certifique-se de fechar a conexão
        if "connection" in locals() and connection:
            connection.close()

@app.route("/ppp/", methods=["GET"])
@app.route("/ppp", methods=["GET"])
def ppp():
    return render_template("ppp.html")


# @app.route("/ppp_je/", methods=["GET"])
# @app.route("/ppp_je", methods=["GET"])
# def ppp_je():
#     return render_template("ppp_je.html")


@app.route("/ppp_tr/", methods=["GET"])
@app.route("/ppp_tr", methods=["GET"])
def ppp_tr():
    return render_template("ppp_tr.html")


@app.route("/seplan_or/", methods=["GET"])
@app.route("/seplan_or", methods=["GET"])
def seplan_or():
    return render_template("seplan_or.html")


@app.route("/", methods=["GET"])
def index():
    return "Extrator DGT"
